from openpyxl import Workbook, load_workbook

book = Workbook() 
sheet = book.active
sheet.title = "Types"

FILENAME_PREFIX = "checking filename: " 
IDENTIFIER_PREFIX = "can't infer type of identifier: "

sheet.cell(row=1, column=1).value =  "Filename"
sheet.cell(row=1, column=2).value =  "File Type"
sheet.cell(row=1, column=3).value =  "Identifier"

def remove_prefix(text, prefix):
	return text[len(prefix):]


f = open("out.txt", encoding = 'utf-8', mode = "r")
contents = f.readlines()

rowNum = 2
for line in contents:
	if line.startswith(FILENAME_PREFIX):
		val = remove_prefix(line, FILENAME_PREFIX)
		rowNum += 1
		sheet.cell(row=rowNum, column=1).value = val
		index = val.rfind(".")
		filetype =  val[index+1:]
		sheet.cell(row=rowNum, column=2).value = filetype
	elif line.startswith(IDENTIFIER_PREFIX):
		rowNum += 1
		val = remove_prefix(line, IDENTIFIER_PREFIX)
		sheet.cell(row=rowNum, column=3).value = val

book.save("cantInfer.xlsx")


