import os
from openpyxl import Workbook, load_workbook

ANY_TOKEN = "$any$"
NULL_TOKEN = "O"

book = Workbook() 
sheet = book.active
sheet.title = "UserAnnotations"

sheet.cell(row=1, column=1).value =  "Filename"
sheet.cell(row=1, column=2).value =  "Precision"
sheet.cell(row=1, column=3).value =  "Recall"


file_itr = 2
for root, dirs, files in os.walk("data/Repos-cleaned"):
	for file in files:
		if file.endswith(".ttokens"):
			filename_y = os.path.join(root, file);
			filename_yhat = os.path.join(root,file) + ".gold"
			f_y = open(filename_y, encoding = 'utf-8', mode = "r")
			f_yhat = open(filename_yhat, encoding = 'utf-8', mode = "r")
			matches = [];
			
			for line in f_y:
				for word in line.split(): #each word is a type
					matches.append([word, "_"])

			i = 0
			for line in f_yhat:
				for word in line.split(): 
					matches[i][1] = word
					i += 1

			
			cleaned_matches = [];
			num_yhat_types = 0
			num_y_types = 0
			for match in matches:
				if not match[0] == NULL_TOKEN:
					num_y_types += 1
					cleaned_matches.append(match)

				if not match[1] == NULL_TOKEN:
					num_yhat_types += 1
			


			num_y_hat_correct = 0
			for match in cleaned_matches:
				if match[0] == match[1]:
					num_y_hat_correct += 1
				elif match[1] == ANY_TOKEN:
					num_y_hat_correct += .5


			try:
				precision = float(num_y_hat_correct) / float(num_yhat_types)
			except:
				precision = 0

			try:
				recall = float(num_y_hat_correct) / float(num_y_types)
			except:
				recall = 0

			sheet.cell(row=file_itr, column=1).value = filename_y 
			sheet.cell(row=file_itr, column=2).value = precision 
			sheet.cell(row=file_itr, column=3).value = recall 
			
#			print(filename_y, "precision", precision, "recall", recall)

			file_itr += 1

book.save("results.xlsx")

